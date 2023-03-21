import pandas as pd 
import openpyxl
from openpyxl import Workbook 
from openpyxl.styles import Font, Alignment 
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas



# Leer el archivo CSV
df = pd.read_csv('ReporteOAGUS_20230312.csv')

# Crear un nuevo libro de trabajo de Excel
wb = Workbook()

# Seleccionar la hoja activa
ws = wb.active


# Establecer el ancho de las columnas
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 12 
ws.column_dimensions['C'].width = 5     #COLUMNA 1
ws.column_dimensions['D'].width = 8     #COLUMNA 2
ws.column_dimensions['E'].width = 5     #COLUMNA 3
ws.column_dimensions['G'].width = 10    #COLUMNA 4
ws.column_dimensions['H'].width = 5    #COLUMNA 5
ws.column_dimensions['I'].width = 5    #COLUMNA 6
ws.column_dimensions['J'].width = 5     #COLUMNA 7
ws.column_dimensions['K'].width = 5     #COLUMNA 8
ws.column_dimensions['L'].width = 5     #COLUMNA 9
ws.column_dimensions['M'].width = 5     #COLUMNA 10
ws.column_dimensions['N'].width = 6     #COLUMNA 11
ws.column_dimensions['O'].width = 8    #COLUMNA 12
ws.column_dimensions['P'].width = 10    #COLUMNA 13

# Establecer la altura deseada en la fila y columna especificada
ws.row_dimensions[1].height = 20


# Establecer el estilo de fuente y alineación
header_font = Font(name='Calibri', size=6, bold=True)
header_alignment = Alignment(horizontal='center', vertical='center')
cell_font = Font(name='Calibri', size=6)
cell_alignment = Alignment(horizontal='left', vertical='center')


# Crear una lista con los nombres de las cabeceras
cabeceras = ['CA', 'Market', 'Ind AM', 'Region', 'Month', 'Chg', 'Prev Ops', 'New Ops', 'Ops Chg', 'Prev Seat', 'New Seat', 'Sea Chg', '%_Seat Chg']

headers = cabeceras                             #Indico el número de columna a iniciar las cabeceras.
for col_num, header_title in enumerate(headers, 3):
    cell = ws.cell(row=1, column=col_num, value=header_title)
    cell.font = header_font
    cell.alignment = header_alignment

# Justifica el texto de los encabezados
for cell in ws[1]:
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    

# Escribir los datos                
for row_num, row_data in enumerate(df.values, 2):
    for col_num, cell_value in enumerate(row_data, 2):
        cell = ws.cell(row=row_num, column=col_num, value=cell_value)
        cell.font = cell_font
        cell.alignment = cell_alignment
    if str(ws.cell(row=row_num, column=4).value) == 'nan':
        ws.delete_rows(row_num)









# Cambiar el color de la celda A1 a rojo

#fill = PatternFill(start_color='808080', end_color='FFFFFF', fill_type='solid')

# Iterar sobre todas las filas y aplicar el formato de relleno al patrón especificado
#for row in ws.iter_rows():
 #   if row[0].value == 'TOTAL':
  #      cell_font = Font(bold = True)
   #     for cell in row:
    #        cell.fill = fill

#color1 = 'BFBFBF' COLOR DE LA CABECERA


# Seleccionar las celdas que se van a ajustar
#cell_range = ws['I1:M1']

# Ajustar el ancho de las columnas para que el texto quepa
#for row in cell_range:
 #   for cell in row:
  #      ws.column_dimensions[cell.column_letter].width = len(str(cell.value))


# definir los colores para los renglones
color1 = 'F2F2F2'
color2 = 'FFFFFF'
change = False


# cambiar el formato de color de los renglones 
for idx, row in enumerate(ws.iter_rows(),1):
    if ((str(ws.cell(row=idx, column=3).value) == 'None') or (str(ws.cell(row=idx, column=4).value) == 'TOTAL')):
        fill = PatternFill(start_color=color2, end_color=color2, fill_type='solid')
        change = True              
    elif change:
        fill = PatternFill(start_color=color1, end_color=color1, fill_type='solid')
        change = False                    
    else:
        fill = PatternFill(start_color=color2, end_color=color2, fill_type='solid')
        change = True
    for cell in row:
        cell.fill = fill
        
# Poner el renglon de totales en negritas
for idx, row in enumerate(ws.iter_rows(),1):
    if str(ws.cell(row=idx, column=4).value) == 'TOTAL':
        for cell in row:
            cell.font = Font(bold=True, name='Calibri', size=6)

# Establecer color de relleno para los encabezados
fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')

# Justifica el texto de los encabezados
for cell in ws[1]:
    cell.fill = fill    #Agrego el color de relleno en el renglon 1


# Buscar la celda que contiene la palabra "TOTAL"
#for fila in ws.rows:
 #   for celda in fila:
  #      if celda.value == 'TOTAL':
            # Obtener la fila donde se encuentra la celda
   #         fila_total = celda.row
            # Poner el texto en negrita en toda la fila
    #        for celda_en_fila in ws[f'A{fila_total}:Z{fila_total}']:
     #           for celda_individual in celda_en_fila:
      #              celda_individual.font = openpyxl.styles.Font(bold=True)





#Convirtiendo a Decimales
# Selecciona la columna que deseas convertir (por ejemplo, columna A)
columna = ws['O']

# Itera sobre las celdas en la columna y convierte los valores de decimal a porcentaje
for celda in columna:
    if isinstance(celda.value, float):  # verifica si el valor de la celda es un decimal
        celda.value = celda.value * 1  # convierte el valor a porcentaje
        celda.number_format = '0%'  # establece el formato de número de la celda como porcentaje con dos decimales




# Indicar el número de columna que deseas eliminar
num_columna = 1
num_columna2 = 1
num_columna17 = 14
num_columna18 = 15
num_columna19 = 14

# Eliminar la columnas

ws.delete_cols(num_columna)
ws.delete_cols(num_columna2)
ws.delete_cols(num_columna17)
ws.delete_cols(num_columna18)
ws.delete_cols(num_columna19)

# Guardar el archivo
wb.save('archivo.xlsx')
